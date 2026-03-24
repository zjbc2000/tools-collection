#!/usr/bin/env python3
"""
XLS 转 XLSX 工具
读取 .xls 文件，拆分所有合并单元格并向下/向右填充原值，输出 .xlsx 文件

依赖: xlrd>=2.0, openpyxl
注意: xlrd 2.x 的 formatting_info=True 在部分文件上会崩溃，
      本脚本改用 xlrd.compdoc 直接解析 OLE2 流来获取合并区域。

可选功能:
  header_row  指定某行（1-based）作为首行，删除其上所有行；
              该行在合并单元格填充时会在值末尾追加 _0、_1… 保证列名唯一。
"""
import sys
import struct
import xlrd
import xlrd.compdoc as compdoc
import openpyxl
from pathlib import Path
from typing import Union, Optional


def _parse_merged_cells(xls_path: str) -> dict[int, list[tuple]]:
    """
    从 XLS 的 OLE2 流中直接解析 MERGEDCELLS 记录，
    绕过 xlrd formatting_info=True 的崩溃问题。

    Returns:
        {xlrd_sheet_idx: [(rlo, rhi, clo, chi), ...]}
        区间为半开区间，与 xlrd.merged_cells 保持一致
    """
    MERGEDCELLS = 0x00E5
    CONTINUE    = 0x003C
    BOF         = 0x0809

    with open(xls_path, "rb") as f:
        raw = f.read()

    cd = compdoc.CompDoc(raw, logfile=open("/dev/null", "w"))
    stream = cd.get_named_stream("Workbook")

    # bof_count: 0 = 全局工作簿 BOF，1 起为各工作表
    bof_count = -1
    results: dict[int, list] = {}
    pending = False
    pos = 0

    while pos + 4 <= len(stream):
        rec_type, rec_len = struct.unpack_from("<HH", stream, pos)
        rec_data = stream[pos + 4 : pos + 4 + rec_len]
        pos += 4 + rec_len

        if rec_type == BOF:
            bof_count += 1
            pending = False
        elif rec_type == MERGEDCELLS:
            pending = True
            sheet_idx = bof_count - 1  # 转换为 xlrd 工作表下标（0-based）
            _read_merge_block(rec_data, sheet_idx, results)
        elif rec_type == CONTINUE and pending:
            _read_merge_block(rec_data, sheet_idx, results)
        else:
            pending = False

    return results


def _read_merge_block(rec_data: bytes, sheet_idx: int, results: dict):
    """解析单个 MERGEDCELLS / CONTINUE 数据块，追加到 results"""
    count = struct.unpack_from("<H", rec_data, 0)[0]
    lst = results.setdefault(sheet_idx, [])
    for i in range(count):
        rlo, rhi, clo, chi = struct.unpack_from("<HHHH", rec_data, 2 + i * 8)
        lst.append((rlo, rhi + 1, clo, chi + 1))  # 半开区间


def _make_header_unique(row: list) -> list:
    """
    对列表中重复的值追加 _0、_1… 使每个元素唯一。
    遍历顺序从左到右，首次出现的值不加后缀，后续重复才加。
    """
    seen: dict[str, int] = {}
    result = []
    for val in row:
        key = str(val)
        if key not in seen:
            seen[key] = 0
            result.append(val)
        else:
            result.append(f"{key}_{seen[key]}")
            seen[key] += 1
    return result


def xls_to_xlsx(
    xls_path: Union[str, Path],
    output_path: Optional[Union[str, Path]] = None,
    header_row: Optional[int] = None,
) -> Path:
    """
    将 .xls 转换为 .xlsx，合并单元格拆分后用左上角原值填充所有子格

    Args:
        xls_path:   输入的 .xls 文件路径
        output_path: 输出路径，默认同目录同名 .xlsx
        header_row: 指定首行行号（1-based）。该行以上的所有行将被删除；
                    该行中属于合并区域的填充值会追加 _0、_1… 保证列名唯一。

    Returns:
        Path: 生成的 .xlsx 文件路径
    """
    xls_path = Path(xls_path)

    if not xls_path.exists():
        raise FileNotFoundError(f"文件不存在: {xls_path}")
    if xls_path.suffix.lower() != ".xls":
        raise ValueError(f"仅支持 .xls 格式，当前: {xls_path.suffix}")

    if output_path is None:
        output_path = xls_path.parent / f"{xls_path.stem}.xlsx"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

    # 解析合并区域（不依赖 formatting_info）
    merged_map = _parse_merged_cells(str(xls_path))

    # 读取单元格值
    wb_in = xlrd.open_workbook(str(xls_path))
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # 移除默认空 sheet

    for sheet_idx, sheet_name in enumerate(wb_in.sheet_names()):
        ws_in = wb_in.sheet_by_name(sheet_name)
        ws_out = wb_out.create_sheet(title=sheet_name)

        nrows, ncols = ws_in.nrows, ws_in.ncols

        # 原始数据矩阵
        data = [
            [ws_in.cell_value(r, c) for c in range(ncols)]
            for r in range(nrows)
        ]

        # 收集 header_row 所在行被哪些合并区域覆盖（用于后续唯一化）
        # header_row 是 1-based，转为 0-based 矩阵下标
        header_row_idx = (header_row - 1) if header_row else None

        # 用合并区域左上角的值填充整个区域（向下 + 向右）
        # rlo/clo 超出实际数据范围时跳过（XLS 可能包含空行的合并区域）
        for rlo, rhi, clo, chi in merged_map.get(sheet_idx, []):
            if rlo >= nrows or clo >= ncols:
                continue
            value = data[rlo][clo]
            for r in range(rlo, min(rhi, nrows)):
                for c in range(clo, min(chi, ncols)):
                    # header_row 行跳过，稍后单独处理以便追加序号
                    if r != header_row_idx:
                        data[r][c] = value

        # 对 header_row 行单独填充：先收集该行各列最终值，再唯一化
        if header_row_idx is not None and header_row_idx < nrows:
            for rlo, rhi, clo, chi in merged_map.get(sheet_idx, []):
                if rlo >= nrows or clo >= ncols:
                    continue
                # 仅处理覆盖到 header_row_idx 的合并区域
                if not (rlo <= header_row_idx < min(rhi, nrows)):
                    continue
                value = data[rlo][clo]
                for c in range(clo, min(chi, ncols)):
                    data[header_row_idx][c] = value
            # 追加序号使列名唯一
            data[header_row_idx] = _make_header_unique(data[header_row_idx])

        # 截掉 header_row 之上的所有行
        start_row = header_row_idx if header_row_idx is not None else 0
        data = data[start_row:]

        # 写入目标 sheet
        for r_idx, row in enumerate(data, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws_out.cell(row=r_idx, column=c_idx, value=value)

    wb_out.save(str(output_path))
    return output_path


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="XLS 转 XLSX，拆分并填充合并单元格")
    parser.add_argument("input",  help="输入 .xls 文件路径")
    parser.add_argument("output", nargs="?", help="输出 .xlsx 文件路径（可选）")
    parser.add_argument(
        "--header-row", "-r", type=int, default=None,
        metavar="N",
        help="指定第 N 行（1-based）作为首行，删除其上所有行，并对该行合并单元格追加唯一序号",
    )
    args = parser.parse_args()

    result = xls_to_xlsx(args.input, args.output, header_row=args.header_row)
    print(f"已生成: {result}")
